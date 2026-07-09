import io

import openpyxl
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

import crud
import exceptions
import models
import schemas

TEMPLATE_HEADERS = [
    "Nama Produk",
    "Barcode",
    "Kategori",
    "Min Stok",
    "Status Aktif",
    "Satuan 1",
    "Harga Beli 1",
    "Harga Jual 1",
    "Satuan 2",
    "Konversi 2",
    "Harga Beli 2",
    "Harga Jual 2",
    "Satuan 3",
    "Konversi 3",
    "Harga Beli 3",
    "Harga Jual 3",
]

REQUIRED_HEADERS = ["nama produk", "kategori", "satuan 1", "harga beli 1", "harga jual 1"]


def _truthy(value) -> bool:
    return str(value).strip().lower() in {"ya", "yes", "true", "1"}


def _normalize_header(value) -> str:
    return str(value).strip().lower() if value is not None else ""


def _blank(value) -> bool:
    return value is None or str(value).strip() == ""


def _price(value) -> float:
    """Blank price cells default to 0; non-blank non-numeric values raise ValueError."""
    if _blank(value):
        return 0.0
    return float(value)


def integrity_error_message(error: IntegrityError) -> str:
    detail = str(error.orig).lower()
    if "barcode" in detail:
        return "Barcode sudah digunakan"
    if "sku" in detail:
        return "SKU sudah digunakan"
    if "product_units" in detail:
        return "Gagal menyimpan data satuan produk, silakan coba lagi"
    return "Data produk bentrok dengan data lain yang sudah ada"


def build_template_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produk"
    ws.append(TEMPLATE_HEADERS)
    ws.append(
        ["Minyak Goreng", "", "Sembako", 10, "Ya", "PCS", 12000, 15000, "PACK", 24, 280000, 340000, "DUS", 288, 3300000, 4000000]
    )
    ws.append(["Kecap Manis", "", "Sembako", 5, "Ya", "PCS", 8000, 10000, "", "", "", "", "", "", "", ""])
    for col, width in zip(
        "ABCDEFGHIJKLMNOP",
        [22, 14, 14, 10, 12, 10, 12, 12, 10, 10, 12, 12, 10, 10, 12, 12],
    ):
        ws.column_dimensions[col].width = width
    note = wb.create_sheet("Petunjuk")
    note.append(["Petunjuk pengisian:"])
    note.append(["- Satu baris = satu produk. SKU dibuat otomatis oleh sistem, tidak perlu diisi."])
    note.append(["- Satuan 1 selalu dianggap sebagai Satuan Dasar produk (konversi otomatis = 1)."])
    note.append(["- Satuan 2 dan Satuan 3 opsional, untuk produk yang punya lebih dari satu satuan (misal PACK, DUS). Kosongkan semua kolomnya jika tidak dipakai."])
    note.append(["- Konversi 2/3 diisi jumlah Satuan 1 yang setara, misal PACK = 24 berarti 1 PACK = 24 Satuan 1."])
    note.append(["- Kategori dan Satuan harus sudah ada di menu Master sebelum import (dicocokkan berdasarkan nama, tanpa membedakan huruf besar/kecil)."])
    note.append(["- Min Stok dan Status Aktif opsional (default 0 dan Ya jika dikosongkan). Status Aktif diisi Ya atau Tidak."])
    return wb


def _parse_extra_unit(cell, row, suffix: str):
    """Returns (unit_data | None, error_message | None) for an optional Satuan 2/3 slot."""
    unit_name_raw = cell(row, f"satuan {suffix}")
    if _blank(unit_name_raw):
        return None, None
    try:
        conversion = int(cell(row, f"konversi {suffix}"))
    except (TypeError, ValueError):
        return None, f"Konversi satuan {suffix} harus berupa angka"
    try:
        buy_price = _price(cell(row, f"harga beli {suffix}"))
        sell_price = _price(cell(row, f"harga jual {suffix}"))
    except ValueError:
        return None, f"Harga Beli/Harga Jual satuan {suffix} harus berupa angka"
    return {
        "unit_name": str(unit_name_raw).strip(),
        "conversion": conversion,
        "buy_price": buy_price,
        "sell_price": sell_price,
    }, None


def import_products_from_excel(db: Session, content: bytes) -> schemas.ProductImportResult:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise exceptions.ValidationError("File tidak bisa dibaca, pastikan formatnya .xlsx")

    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise exceptions.ValidationError("File Excel kosong")

    header_map = {_normalize_header(h): i for i, h in enumerate(rows[0])}
    missing = [h for h in REQUIRED_HEADERS if h not in header_map]
    if missing:
        raise exceptions.ValidationError(f"Kolom wajib tidak ditemukan: {', '.join(missing)}")

    def cell(row, name):
        idx = header_map.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    total_rows = 0
    created = 0
    errors: list[schemas.ProductImportRowError] = []

    for row_num, row in enumerate(rows[1:], start=2):
        if row is None or all(_blank(c) for c in row):
            continue
        total_rows += 1

        name_raw = cell(row, "nama produk")
        if _blank(name_raw):
            errors.append(schemas.ProductImportRowError(row=row_num, product_name=None, message="Nama produk kosong"))
            continue
        name = str(name_raw).strip()

        category_raw = cell(row, "kategori")
        if _blank(category_raw):
            errors.append(schemas.ProductImportRowError(row=row_num, product_name=name, message="Kategori kosong"))
            continue
        category = (
            db.query(models.Category).filter(func.lower(models.Category.name) == str(category_raw).strip().lower()).first()
        )
        if category is None:
            errors.append(
                schemas.ProductImportRowError(
                    row=row_num, product_name=name, message=f"Kategori '{category_raw}' tidak ditemukan"
                )
            )
            continue

        satuan1_raw = cell(row, "satuan 1")
        if _blank(satuan1_raw):
            errors.append(
                schemas.ProductImportRowError(row=row_num, product_name=name, message="Satuan 1 (satuan dasar) kosong")
            )
            continue
        try:
            buy1 = _price(cell(row, "harga beli 1"))
            sell1 = _price(cell(row, "harga jual 1"))
        except ValueError:
            errors.append(
                schemas.ProductImportRowError(
                    row=row_num, product_name=name, message="Harga Beli 1 / Harga Jual 1 harus berupa angka"
                )
            )
            continue

        units_data = [
            {"unit_name": str(satuan1_raw).strip(), "conversion": 1, "buy_price": buy1, "sell_price": sell1}
        ]

        row_error = None
        for suffix in ("2", "3"):
            extra, err = _parse_extra_unit(cell, row, suffix)
            if err:
                row_error = err
                break
            if extra:
                units_data.append(extra)
        if row_error:
            errors.append(schemas.ProductImportRowError(row=row_num, product_name=name, message=row_error))
            continue

        unit_creates = []
        unit_missing = None
        for u in units_data:
            unit = db.query(models.Unit).filter(func.lower(models.Unit.name) == u["unit_name"].lower()).first()
            if unit is None:
                unit_missing = u["unit_name"]
                break
            unit_creates.append(
                schemas.ProductUnitCreate(
                    unit_id=unit.id, conversion=u["conversion"], buy_price=u["buy_price"], sell_price=u["sell_price"]
                )
            )
        if unit_missing:
            errors.append(
                schemas.ProductImportRowError(row=row_num, product_name=name, message=f"Satuan '{unit_missing}' tidak ditemukan")
            )
            continue

        barcode = cell(row, "barcode")
        min_stock_raw = cell(row, "min stok")
        status_raw = cell(row, "status aktif")

        try:
            product_in = schemas.ProductCreate(
                name=name,
                barcode=str(barcode).strip() if not _blank(barcode) else None,
                category_id=category.id,
                base_unit_id=unit_creates[0].unit_id,
                min_stock=int(min_stock_raw) if not _blank(min_stock_raw) else 0,
                is_active=_truthy(status_raw) if not _blank(status_raw) else True,
                units=unit_creates,
            )
            crud.create_product(db, product_in)
            created += 1
        except exceptions.ValidationError as e:
            db.rollback()
            errors.append(schemas.ProductImportRowError(row=row_num, product_name=name, message=str(e)))
        except IntegrityError as e:
            db.rollback()
            errors.append(
                schemas.ProductImportRowError(row=row_num, product_name=name, message=_integrity_error_message(e))
            )

    return schemas.ProductImportResult(total_rows=total_rows, created=created, errors=errors)
